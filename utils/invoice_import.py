"""
Shared utilities for Telus ServerSideInvoiceImport file generation.

All client processors (Baker Boy, Reser, McCormick, Maple Leaf, Parway, Lactalis)
produce the same 46-column Telus import format: 33 core + 8 extras + 5 trailer.
This module provides the shared constants, helper functions, and export writer.
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Optional
from dataclasses import dataclass, field

# ============================================================
# TELUS IMPORT FORMAT CONSTANTS (shared across all clients)
# ============================================================

EXPORT_HEADERS_CORE = [
    "Broker",
    "MfrCode",
    "DistrName",
    "DistrCode",
    "PONumber",
    "InvoiceNo",
    "InvoiceDate",
    "MfrItemCode",
    "GtinOrUpcCode",
    "ItemDescription",
    "Quantity",
    "UnitPrice",
    "NetAmt",
    "AmtPaid",
    "NetCatchWt",
    "GrossCatchWt",
    "OfficeCode",
    "PODate",
    "ShipDate",
    "ArrvDate",
    "PickupDate",
    "ShipTo",
    "ShipFrom",
    "MfrName",
    "DistItemCode",
    "ShipFromNbr",
    "DistrBillToNum",
    "DistrShipToNum",
    "MfrInvDistCode",
    "MfrInvShipToCode",
    "PriceModeOverride",
    "PriceListNameOverride",
    "MfrRefNum",
]

EXTRA_IMPORT_COLUMNS = [
    "Region",
    "Dist State",
    "Ship to State",
    "Ship to Office",
    "Foodmark Code",
    "Foodmark Ship To",
    "# of EA in Case",
    "Case Conversion",
]

EXPORT_TRAILER_HEADERS = [
    "DupePo",
    "DupeInv",
    "CalcCasesByWeight",
    "IgnoreCredits",
    "User",
]

EXPORT_HEADERS = EXPORT_HEADERS_CORE + EXTRA_IMPORT_COLUMNS + EXPORT_TRAILER_HEADERS

CORE_IDX_QUANTITY = EXPORT_HEADERS_CORE.index("Quantity")
CORE_IDX_DISTRCODE = EXPORT_HEADERS_CORE.index("DistrCode")
CORE_IDX_SHIPTO = EXPORT_HEADERS_CORE.index("ShipTo")
CORE_IDX_INVOICENO = EXPORT_HEADERS_CORE.index("InvoiceNo")
CORE_IDX_PONUMBER = EXPORT_HEADERS_CORE.index("PONumber")


# ============================================================
# DATA CLASSES
# ============================================================

@dataclass
class ProcessResult:
    """Result from a client processor."""
    rows: list[tuple] = field(default_factory=list)
    dates_seen: list[str] = field(default_factory=list)
    reference_misses: dict[str, dict] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    files_processed: list[str] = field(default_factory=list)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def safe(v: Any) -> str:
    """Safely convert any value to a stripped string."""
    if v is None:
        return ""
    return str(v).strip()


def customer_lookup_key(ctx_customer: str) -> str:
    """Normalize customer name for reference lookup (uppercase, single-spaced)."""
    return " ".join(safe(ctx_customer).split()).upper()


def parse_money(v: Any) -> Optional[float]:
    """Parse (1.23), -1.23, $1,234.56 to float. None if empty."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = safe(v)
    if not s:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s or s == "-" or s == ".":
        return None
    try:
        x = float(s)
        return -x if neg else x
    except ValueError:
        return None


def parse_qty(v: Any) -> Optional[float]:
    """Parse quantity value to float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = safe(v).rstrip(".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def format_invoice_date(v: Any) -> str:
    """Format a date value to MM/DD/YYYY string."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%m/%d/%Y")
    s = safe(v)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%m/%d/%Y")
        except ValueError:
            continue
    return s


def normalize_header(v: Any) -> str:
    """Normalize a cell value for header matching."""
    if v is None:
        return ""
    return str(v).replace("\u00a0", " ").strip().lower()


def should_drop_zero_quantity_row(core: tuple, drop_zeros: bool = True) -> bool:
    """True if this row should be omitted (Quantity missing or numerically zero)."""
    if not drop_zeros:
        return False
    if len(core) <= CORE_IDX_QUANTITY:
        return True
    v = core[CORE_IDX_QUANTITY]
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    try:
        return float(v) == 0.0
    except (TypeError, ValueError):
        return False


def extend_export_row(
    base_core: tuple | list,
    user: str = "LisaFMS",
    region: str = "",
    dist_state: str = "",
    ship_to_state: str = "",
    ship_to_office: str = "",
    ea_per_case: str = "",
    case_conversion: str = "",
    dupe_po: str = "Y",
    dupe_inv: str = "Y",
    calc_cases_by_weight: str = "",
    ignore_credits: str = "",
) -> list:
    """Append extra columns + trailer to a 33-element core row."""
    if len(base_core) != len(EXPORT_HEADERS_CORE):
        raise ValueError(
            f"Expected {len(EXPORT_HEADERS_CORE)} core columns, got {len(base_core)}"
        )
    dist_code = safe(base_core[CORE_IDX_DISTRCODE])
    ship_to = safe(base_core[CORE_IDX_SHIPTO])
    extra = [
        region,
        dist_state,
        ship_to_state,
        ship_to_office,
        dist_code,       # Foodmark Code
        ship_to,         # Foodmark Ship To
        ea_per_case,
        case_conversion,
    ]
    trailer = [dupe_po, dupe_inv, calc_cases_by_weight, ignore_credits, user]
    return list(base_core) + extra + trailer


def write_export_bytes(
    rows: list[tuple],
    user: str = "LisaFMS",
    drop_zeros: bool = True,
    sort_by_invoice: bool = True,
    **extend_kwargs,
) -> tuple[bytes, int]:
    """
    Write Telus import Excel file to bytes (for Streamlit download).
    
    Returns (xlsx_bytes, rows_written).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    # Filter zero-qty rows
    filtered = [r for r in rows if not should_drop_zero_quantity_row(r, drop_zeros)]
    if not filtered:
        return b"", 0

    # Sort by InvoiceNo then PONumber
    if sort_by_invoice:
        filtered = sorted(
            filtered,
            key=lambda r: (str(r[CORE_IDX_INVOICENO] or "").strip(), str(r[CORE_IDX_PONUMBER] or "").strip())
        )

    # Extend rows with extras + trailer
    full_rows = [extend_export_row(r, user=user, **extend_kwargs) for r in filtered]

    # Write to openpyxl workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Data"
    for i, h in enumerate(EXPORT_HEADERS, 1):
        c = ws.cell(1, i, value=h)
        c.font = Font(bold=True)
    for ri, row in enumerate(full_rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, value=val)

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), len(full_rows)


def coverage_label_from_dates(dates: list[str]) -> str:
    """Generate a month label from parsed dates (e.g. 'January_2026')."""
    if not dates:
        return datetime.now().strftime("%B_%Y")
    from collections import Counter
    parsed = []
    for d in dates:
        try:
            parsed.append(datetime.strptime(d, "%m/%d/%Y"))
        except ValueError:
            continue
    if not parsed:
        return datetime.now().strftime("%B_%Y")
    ym = [(p.year, p.month) for p in parsed]
    (y, mo), _ = Counter(ym).most_common(1)[0]
    return f"{datetime(y, mo, 1).strftime('%B')}_{y}"
