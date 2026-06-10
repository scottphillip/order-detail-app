"""
Baker Boy commission processor for Streamlit.

Adapts the desktop script's parsing logic to work with Streamlit UploadedFile objects
and Snowflake-stored reference data (DataFrame instead of local Excel).

Supports 3 source layouts:
- Legacy: merged "No. 12345" cells
- JF/Dakota column-C: Doc. Type in col C, item codes in col D
- Aptean Trade Statement: new ERP format (post May 2026)
"""
from __future__ import annotations

import io
import re
import zipfile
import tempfile
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook

from utils.invoice_import import (
    ProcessResult,
    safe,
    customer_lookup_key,
    parse_money,
    parse_qty,
    format_invoice_date,
    normalize_header,
)

STATIC_MFRCODE = "BAKERBOY"
MAX_HEADER_SCAN_ROWS = 35

NO_ITEM_RE = re.compile(r"No\.\s*([\d.]+)", re.I)
SIP_RE = re.compile(r"SIP\s*[\d]+", re.I)
DOC_COL_ITEM_LEAK = re.compile(r"^No\.\s*[\d.]+\s*$", re.I)
REPORT_FOOTER_MARKERS = (
    "report totals", "total invoice amount", "total credit memo amount",
    "total adjustment amount", "total amount", "commission to be paid",
)
APTEAN_SUBTOTAL_RE = re.compile(r"subtotal\s+for\s+(\S+)\s+(.*?)\s*:\s*$", re.I)


def _build_ref_dict(refs_df: pd.DataFrame) -> dict[str, tuple[str, str, str, str]]:
    """Convert reference DataFrame to lookup dict: key -> (dist, ship, pm, pl)."""
    ref = {}
    if refs_df.empty:
        return ref
    for _, row in refs_df.iterrows():
        k = str(row.get("LOOKUP_KEY", "")).strip()
        if not k:
            continue
        ref[k] = (
            str(row.get("DIST_CODE", "") or "").strip(),
            str(row.get("SHIP_TO", "") or "").strip(),
            str(row.get("PRICE_MODE", "") or "").strip(),
            str(row.get("PRICE_LIST", "") or "").strip(),
        )
    return ref


def cell_merged(ws, row: int, col: int) -> Any:
    v = ws.cell(row=row, column=col).value
    if v is not None and safe(str(v)) != "":
        return v
    try:
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return ws.cell(rng.min_row, rng.min_col).value
    except Exception:
        pass
    return v


def is_report_footer_row(ws, r: int, scan_cols: int = 12) -> bool:
    parts = []
    for c in range(1, min(ws.max_column + 1, scan_cols + 1)):
        v = safe(str(cell_merged(ws, r, c))).lower()
        if v:
            parts.append(v)
    blob = " ".join(parts)
    return any(m in blob for m in REPORT_FOOTER_MARKERS)


def map_header_row(ws, hr: int) -> dict[str, int]:
    keys = {}
    for c in range(1, ws.max_column + 1):
        h = normalize_header(ws.cell(hr, c).value)
        if not h:
            continue
        if "external" in h and "doc" in h:
            keys.setdefault("ext", c)
        hn = h.replace(".", "").replace(" ", "")
        if "documentno" in hn or re.match(r"^document\s*no\.?$", h.strip()):
            if "external" not in h:
                keys.setdefault("doc", c)
        if "customer" in h and "name" in h:
            keys.setdefault("cust", c)
        if "posting" in h and "date" in h:
            keys.setdefault("post", c)
        if h == "doc. type" or (h.startswith("doc") and "type" in h):
            keys.setdefault("doctype", c)
        if "commission" in h and "amount" in h:
            keys.setdefault("comm", c)
    return keys


def _row_has_sip_and_po(ws, r, doc_col, ext_col):
    d = safe(str(cell_merged(ws, r, doc_col)))
    e = safe(str(cell_merged(ws, r, ext_col)))
    if e.startswith("=") or d.startswith("="):
        return False
    return bool(SIP_RE.search(d) or (len(d) >= 8 and re.search(r"[A-Za-z]{2,}\d", d))) and bool(
        e and re.search(r"\d", e)
    )


def find_best_header_row(ws) -> tuple[int, dict]:
    best = None
    for hr in range(1, min(MAX_HEADER_SCAN_ROWS + 15, ws.max_row + 1)):
        hdr = map_header_row(ws, hr)
        dc, ec = hdr.get("doc"), hdr.get("ext")
        if not dc or not ec:
            continue
        sip_like = 0
        lim = min(hr + 120, ws.max_row + 1)
        for r in range(hr + 1, lim):
            if is_report_footer_row(ws, r):
                break
            if _row_has_sip_and_po(ws, r, dc, ec):
                sip_like += 1
        if sip_like > 0:
            if best is None or sip_like > best[0]:
                best = (sip_like, hr, hdr)
    if best:
        return best[1], map_header_row(ws, best[1])
    # fallback
    for r in range(1, min(MAX_HEADER_SCAN_ROWS, ws.max_row + 1)):
        texts = [normalize_header(ws.cell(r, c).value) for c in range(1, min(ws.max_column + 1, 20))]
        joined = " ".join(texts)
        if "document" in joined and "external" in joined:
            return r, map_header_row(ws, r)
    return 12, map_header_row(ws, 12)


def extract_item_from_row(ws, r: int) -> Optional[str]:
    for c in range(1, min(ws.max_column + 1, 22)):
        v = cell_merged(ws, r, c)
        if v is None:
            continue
        m = NO_ITEM_RE.search(str(v))
        if m:
            return m.group(1).replace(".0", "").rstrip(".")
    return None


def is_summary_invoice_row(ws, r, doc_col):
    v = safe(cell_merged(ws, r, doc_col)).upper()
    if not v or v == "QUANTITY":
        return False
    if SIP_RE.search(v) or v.startswith("SIP"):
        return True
    if len(v) > 5 and re.match(r"^[A-Z]{2,}\d", v):
        return True
    return False


def _record_miss(tracker, ctx_customer, ref):
    k = customer_lookup_key(ctx_customer)
    if k not in tracker:
        reason = "Not in reference table" if k not in ref else "In reference but Dist/Ship blank"
        tracker[k] = {
            "customer_name": safe(ctx_customer) or "<blank>",
            "lookup_key": k,
            "line_count": 0,
            "reason": reason,
        }
    tracker[k]["line_count"] += 1


def scan_sheet_legacy(ws, ref: dict) -> tuple[list[tuple], list[str], dict]:
    """Parse legacy/standard layout sheet."""
    hr, hdr = find_best_header_row(ws)
    ext_col = hdr.get("ext")
    doc_col = hdr.get("doc")
    cust_col = hdr.get("cust")
    post_col = hdr.get("post")
    comm_col = hdr.get("comm") or 11

    if not doc_col or not ext_col:
        return [], [], {}

    qty_col = cust_col or 7
    ctx_po = ctx_inv = ctx_date = ctx_customer = ""
    rows_out, dates_seen, misses = [], [], {}

    for r in range(hr + 1, ws.max_row + 1):
        if is_report_footer_row(ws, r):
            break
        doc_s = safe(cell_merged(ws, r, doc_col))
        if doc_s and DOC_COL_ITEM_LEAK.match(doc_s):
            continue

        item = extract_item_from_row(ws, r)
        if item and ctx_inv:
            q = parse_qty(cell_merged(ws, r, qty_col))
            amt = parse_money(cell_merged(ws, r, comm_col))
            if q is None:
                q = 0.0
            if amt is not None and amt < 0:
                q = -abs(q) if q >= 0 else q
            k = customer_lookup_key(ctx_customer)
            dist, ship, pm, pl = ref.get(k, ("", "", "", ""))
            dist = dist or ship
            if not safe(dist):
                _record_miss(misses, ctx_customer, ref)
            rows_out.append((
                "", STATIC_MFRCODE, "", dist, ctx_po or ctx_inv, ctx_inv, ctx_date,
                item, "", "", q, "", "", amt, "", "", "", "", "", "", "",
                ship, "", "", "", "", "", "", "", "", pm, pl, "",
            ))
            continue

        ext_s = safe(cell_merged(ws, r, ext_col))
        is_summary = False
        if ext_s and doc_s and doc_s.lower() != "quantity":
            if SIP_RE.search(doc_s) or (len(doc_s) >= 6 and re.search(r"[A-Za-z]{2,}\d", doc_s) and "no." not in doc_s.lower()):
                is_summary = True
        if not is_summary and ext_s and is_summary_invoice_row(ws, r, doc_col):
            is_summary = True
        if is_summary:
            ctx_po = ext_s
            ctx_inv = doc_s
            ctx_date = format_invoice_date(cell_merged(ws, r, post_col)) if post_col else ""
            ctx_customer = safe(cell_merged(ws, r, cust_col)) if cust_col else ""
            if ctx_date:
                dates_seen.append(ctx_date)

    return rows_out, dates_seen, misses


def jf_header_row(ws) -> int:
    for r in range(1, min(MAX_HEADER_SCAN_ROWS, (ws.max_row or 0)) + 1):
        c = normalize_header(ws.cell(r, 3).value)
        if "doc" in c and "type" in c:
            return r
    return 0


def looks_like_jf_layout(ws) -> bool:
    return jf_header_row(ws) > 0


def scan_sheet_jf(ws, ref: dict) -> tuple[list[tuple], list[str], dict]:
    """Parse JF/Dakota column-C layout."""
    rows_out, dates_seen, misses = [], [], {}
    ctx_po = ctx_inv = ctx_date = ctx_customer = ""
    in_detail_block = False

    hr = jf_header_row(ws)
    data_start = (hr + 1) if hr else 14
    JF_COL_C, JF_COL_D, JF_COL_F, JF_COL_G, JF_COL_H, JF_COL_K = 3, 4, 6, 7, 8, 11

    for r in range(data_start, (ws.max_row or 0) + 1):
        a = safe(str(cell_merged(ws, r, 1))).upper()
        if "REPORT TOTALS" in a:
            break
        c_val = safe(str(cell_merged(ws, r, JF_COL_C))).lower().strip()
        if any(m in c_val for m in ("total invoice", "total credit", "commission to be")):
            break

        d_val = safe(cell_merged(ws, r, JF_COL_D))

        if c_val in ("invoice", "credit"):
            ctx_po = d_val
            ctx_inv = safe(cell_merged(ws, r, JF_COL_F))
            ctx_customer = safe(cell_merged(ws, r, JF_COL_G))
            ctx_date = format_invoice_date(cell_merged(ws, r, JF_COL_H))
            if ctx_date:
                dates_seen.append(ctx_date)
            in_detail_block = False
            continue

        if c_val in ("no.", "no"):
            in_detail_block = True
        elif not in_detail_block:
            continue

        if not d_val or not re.fullmatch(r"\d+(?:\.\d+)?", d_val) or "." in d_val or len(d_val) > 10 or len(d_val) < 2:
            if not d_val and in_detail_block:
                in_detail_block = False
            continue

        if not ctx_inv:
            continue

        qty = parse_qty(cell_merged(ws, r, JF_COL_H))
        amt = parse_money(cell_merged(ws, r, JF_COL_K))
        if qty is None and amt is None:
            continue
        if qty is None:
            qty = 0.0
        if amt is not None and amt < 0:
            qty = -abs(qty) if qty >= 0 else qty

        k = customer_lookup_key(ctx_customer)
        dist, ship, pm, pl = ref.get(k, ("", "", "", ""))
        dist = dist or ship
        if not safe(dist):
            _record_miss(misses, ctx_customer, ref)

        rows_out.append((
            "", STATIC_MFRCODE, "", dist, ctx_po or ctx_inv, ctx_inv, ctx_date,
            d_val, "", "", qty, "", "", amt, "", "", "", "", "", "", "",
            ship, "", "", "", "", "", "", "", "", pm, pl, "",
        ))

    return rows_out, dates_seen, misses


def looks_like_aptean(ws) -> bool:
    for r in range(1, min(40, (ws.max_row or 0)) + 1):
        joined = " ".join(normalize_header(ws.cell(r, c).value) for c in range(1, min((ws.max_column or 0) + 1, 16)))
        if ("cust. po no" in joined or "cust po no" in joined) and ("eligible qty" in joined or "amount granted" in joined or "item no" in joined):
            return True
    return False


def scan_sheet_aptean(ws, ref: dict) -> tuple[list[tuple], list[str], dict]:
    """Parse Aptean Trade Statement layout."""
    rows_out, dates_seen, misses = [], [], {}
    maxr = ws.max_row or 0
    maxc = ws.max_column or 0

    # Find header
    cols = {}
    hr = 0
    for r in range(1, min(40, maxr) + 1):
        for c in range(1, maxc + 1):
            h = normalize_header(ws.cell(r, c).value)
            if not h:
                continue
            if "posting date" in h:
                cols.setdefault("date", c)
            if "document no" in h and "external" not in h:
                cols.setdefault("inv", c)
            if "cust" in h and "po no" in h:
                cols.setdefault("po", c)
            if "item no" in h and "description" not in h:
                cols.setdefault("item", c)
            if "eligible qty" in h:
                cols.setdefault("qty", c)
            if "amount granted" in h:
                cols.setdefault("amt", c)
        if "po" in cols and "item" in cols:
            hr = r
            break

    if not hr:
        return [], [], {}

    # Build source code -> customer name from subtotal rows
    code2name = {}
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            v = cell_merged(ws, r, c)
            if v is None:
                continue
            m = APTEAN_SUBTOTAL_RE.search(str(v))
            if m:
                code, name = m.group(1).strip(), m.group(2).strip()
                if code and name:
                    code2name.setdefault(code, name)
                break

    current_code = ""
    for r in range(1, maxr + 1):
        a_txt = safe(cell_merged(ws, r, 1)).lower()
        if a_txt.startswith("source"):
            for c in range(2, maxc + 1):
                cv = safe(cell_merged(ws, r, c))
                if cv and "source" not in cv.lower():
                    current_code = cv
                    break
            continue

        item = safe(cell_merged(ws, r, cols.get("item", 1)))
        if not item or " " in item or item.lower() in ("item no.", "item no"):
            continue

        inv = safe(cell_merged(ws, r, cols.get("inv", 1))) if "inv" in cols else ""
        date_s = format_invoice_date(cell_merged(ws, r, cols.get("date", 1))) if "date" in cols else ""
        if not (date_s or inv):
            continue

        po = safe(cell_merged(ws, r, cols.get("po", 1))) if "po" in cols else ""
        qty = parse_qty(cell_merged(ws, r, cols.get("qty", 1))) if "qty" in cols else None
        amt = parse_money(cell_merged(ws, r, cols.get("amt", 1))) if "amt" in cols else None
        if qty is None and amt is None:
            continue
        if qty is None:
            qty = 0.0
        if amt is not None and amt < 0:
            qty = -abs(qty) if qty >= 0 else qty

        customer = code2name.get(current_code, "")
        k = customer_lookup_key(customer)
        dist, ship, pm, pl = ref.get(k, ("", "", "", ""))
        dist = dist or ship
        if not safe(dist):
            _record_miss(misses, customer, ref)
        if date_s:
            dates_seen.append(date_s)

        rows_out.append((
            "", STATIC_MFRCODE, "", dist, po or inv, inv, date_s,
            item, "", "", qty, "", "", amt, "", "", "", "", "", "", "",
            ship, "", "", "", "", "", "", "", "", pm, pl, "",
        ))

    return rows_out, dates_seen, misses


def ordered_sheet_names(names: list[str]) -> list[str]:
    def score(sn):
        s = sn.lower()
        if "detail" in s and "commission" in s:
            return 100
        if "detail" in s:
            return 80
        if "commission" in s:
            return 70
        if "copy" in s or s.startswith("jf "):
            return 5
        return 40
    return sorted(names, key=lambda n: -score(n))


def process_one_workbook(file_bytes: bytes, filename: str, ref: dict) -> tuple[list[tuple], list[str], dict]:
    """Process a single workbook from bytes."""
    best_rows, best_dates, best_misses = [], [], {}

    for data_only in (True, False):
        wb = load_workbook(io.BytesIO(file_bytes), data_only=data_only, read_only=False, keep_vba=False)
        try:
            names = ordered_sheet_names(list(wb.sheetnames))
            pass_rows, pass_dates, pass_misses = [], [], {}

            for sn in names:
                ws = wb[sn]
                if (ws.max_row or 0) < 3:
                    continue
                if looks_like_aptean(ws):
                    rows, ds, ms = scan_sheet_aptean(ws, ref)
                elif looks_like_jf_layout(ws):
                    rows, ds, ms = scan_sheet_jf(ws, ref)
                else:
                    rows, ds, ms = scan_sheet_legacy(ws, ref)
                pass_rows.extend(rows)
                pass_dates.extend(ds)
                for k, v in ms.items():
                    if k not in pass_misses:
                        pass_misses[k] = v
                    else:
                        pass_misses[k]["line_count"] += v["line_count"]
        finally:
            wb.close()

        if len(pass_rows) > len(best_rows):
            best_rows, best_dates, best_misses = pass_rows, pass_dates, pass_misses

    return best_rows, best_dates, best_misses


def process_baker_boy(uploaded_files, refs_df: pd.DataFrame) -> ProcessResult:
    """
    Main entry point for Baker Boy processing.
    
    Args:
        uploaded_files: list of Streamlit UploadedFile objects
        refs_df: DataFrame with columns LOOKUP_KEY, DIST_CODE, SHIP_TO, PRICE_MODE, PRICE_LIST
    
    Returns:
        ProcessResult with rows, dates, reference misses, warnings
    """
    ref = _build_ref_dict(refs_df)
    result = ProcessResult()

    for uf in uploaded_files:
        file_bytes = uf.read()
        filename = uf.name

        if filename.lower().endswith(".zip"):
            # Extract zip and process each workbook inside
            with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
                for info in zf.infolist():
                    if info.filename.endswith("/"):
                        continue
                    if not any(info.filename.lower().endswith(ext) for ext in (".xlsx", ".xlsm")):
                        continue
                    if info.filename.startswith("~$"):
                        continue
                    wb_bytes = zf.read(info)
                    try:
                        rows, dates, misses = process_one_workbook(wb_bytes, info.filename, ref)
                        result.rows.extend(rows)
                        result.dates_seen.extend(dates)
                        for k, v in misses.items():
                            if k not in result.reference_misses:
                                result.reference_misses[k] = v
                            else:
                                result.reference_misses[k]["line_count"] += v["line_count"]
                        result.files_processed.append(info.filename)
                        if not rows:
                            result.warnings.append(f"{info.filename}: No detail lines found")
                    except Exception as e:
                        result.warnings.append(f"{info.filename}: Error - {str(e)[:200]}")
        else:
            # Direct workbook
            try:
                rows, dates, misses = process_one_workbook(file_bytes, filename, ref)
                result.rows.extend(rows)
                result.dates_seen.extend(dates)
                for k, v in misses.items():
                    if k not in result.reference_misses:
                        result.reference_misses[k] = v
                    else:
                        result.reference_misses[k]["line_count"] += v["line_count"]
                result.files_processed.append(filename)
                if not rows:
                    result.warnings.append(f"{filename}: No detail lines found")
            except Exception as e:
                result.warnings.append(f"{filename}: Error - {str(e)[:200]}")

    return result
